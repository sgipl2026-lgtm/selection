#!/usr/bin/env python3
"""
update_matches.py
-----------------
1. Reads data.json
2. Marks matches as "Complete" if their IST datetime + 3h buffer has passed
3. Injects updated SQUADS, MATCHES, and USER_PINS into index.html (sentinel-based)
4. Regenerates user_list.xlsx from the users array in data.json
5. Commits all changed files via git (done by GitHub Actions)

Dependencies: openpyxl  (pip install openpyxl)
Run:          python scripts/update_matches.py
"""

import json
import re
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed - user_list.xlsx will NOT be regenerated.")

REPO_ROOT  = Path(__file__).resolve().parent.parent
DATA_FILE  = REPO_ROOT / "data.json"
INDEX_FILE = REPO_ROOT / "index.html"
XLSX_FILE  = REPO_ROOT / "user_list.xlsx"

IST = timezone(timedelta(hours=5, minutes=30))

SQUADS_START  = "// @@BEGIN_SQUADS@@"
SQUADS_END    = "// @@END_SQUADS@@"
MATCHES_START = "// @@BEGIN_MATCHES@@"
MATCHES_END   = "// @@END_MATCHES@@"
USERS_START   = "// @@BEGIN_USERS@@"
USERS_END     = "// @@END_USERS@@"

APP_URL = "https://sgipl2026-lgtm.github.io/selection/"


def load_data():
    with open(DATA_FILE, encoding="utf-8") as f:
        return json.load(f)


def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print("data.json saved.")


def update_match_statuses(data):
    now_ist = datetime.now(IST)
    changed = False
    for match in data["matches"]:
        if match["status"] == "Complete":
            continue
        try:
            match_dt = datetime.strptime(
                f"{match['date']} {match['time']}", "%Y-%m-%d %H:%M"
            ).replace(tzinfo=IST)
        except ValueError:
            print(f"  WARNING: Bad datetime for {match['id']} - skipping")
            continue
        if now_ist >= match_dt + timedelta(hours=3):
            print(f"  Marking '{match['name']}' Complete")
            match["status"] = "Complete"
            changed = True
    return data, changed


def squads_to_js(squads):
    lines = ["const SQUADS = {"]
    entries = []
    for team, players in squads.items():
        ps = ",".join(f'{{n:"{p["n"]}",c:"{p["c"]}"}}' for p in players)
        entries.append(f"  {team}: [{ps}]")
    lines.append(",\n".join(entries) + ",")
    lines.append("};")
    return "\n".join(lines)


def matches_to_js(matches):
    lines = ["const MATCHES = ["]
    entries = []
    for m in matches:
        entries.append(
            f'  {{id:"{m["id"]}",name:"{m["name"]}",team1:"{m["team1"]}",team2:"{m["team2"]}",'
            f'date:"{m["date"]}",time:"{m["time"]}",status:"{m["status"]}"}}'
        )
    lines.append(",\n".join(entries))
    lines.append("];")
    return "\n".join(lines)


def users_to_js(users):
    lines = [
        '// Format: "Name": { id:"UXXXXXXXX", pin:"NNNN", email:"..." }',
        "const USER_PINS = {"
    ]
    entries = []
    for u in users:
        name  = u["name"].replace('"', '\\"')
        uid   = u["id"]
        pin   = u["pin"]
        email = u.get("email", "").replace('"', '\\"')
        entries.append(f'  "{name}": {{ id:"{uid}", pin:"{pin}", email:"{email}" }}')
    lines.append(",\n".join(entries) + ",")
    lines.append("};")
    return "\n".join(lines)


def inject_block(html, start_marker, end_marker, new_inner):
    pattern = re.compile(
        rf"({re.escape(start_marker)}\n).*?(\n{re.escape(end_marker)})",
        re.DOTALL,
    )
    result, count = pattern.subn(rf"\g<1>{new_inner}\g<2>", html)
    if count == 0:
        raise ValueError(f"Sentinel not found: {start_marker}")
    return result


def inject_into_html(squads_js, matches_js, users_js):
    html = INDEX_FILE.read_text(encoding="utf-8")
    original = html
    html = inject_block(html, SQUADS_START,  SQUADS_END,  squads_js)
    html = inject_block(html, MATCHES_START, MATCHES_END, matches_js)
    html = inject_block(html, USERS_START,   USERS_END,   users_js)
    if html == original:
        print("index.html already up to date.")
        return False
    INDEX_FILE.write_text(html, encoding="utf-8")
    print("index.html updated.")
    return True


def generate_xlsx(users):
    if not HAS_OPENPYXL:
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    HEADER_FILL  = PatternFill("solid", fgColor="1E2D45")
    HEADER_FONT  = Font(bold=True, color="F5A623", size=11)
    ROW_FONT     = Font(color="E8EAF0", size=10)
    ALT_FILL     = PatternFill("solid", fgColor="111827")
    NORMAL_FILL  = PatternFill("solid", fgColor="1A2235")
    CTR          = Alignment(horizontal="center", vertical="center")
    LEFT         = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="1E2D45")
    BDR          = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["User ID", "User Name", "User Email ID", "User PIN", "User URL"]
    col_widths = [16, 18, 34, 12, 52]

    for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CTR; cell.border = BDR
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for ri, user in enumerate(users, 2):
        fill   = NORMAL_FILL if ri % 2 == 0 else ALT_FILL
        values = [user.get("id",""), user.get("name",""), user.get("email",""),
                  user.get("pin",""), APP_URL]
        aligns = [CTR, LEFT, LEFT, CTR, LEFT]
        for ci, (val, aln) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = ROW_FONT; cell.fill = fill
            cell.alignment = aln; cell.border = BDR
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A2"
    wb.save(XLSX_FILE)
    print(f"user_list.xlsx saved ({len(users)} users).")
    return True


def main():
    print("-" * 60)
    print(f"update_matches.py  --  {datetime.now(IST).strftime('%Y-%m-%d %H:%M IST')}")
    print("-" * 60)

    data = load_data()
    data, data_changed = update_match_statuses(data)

    squads_js  = squads_to_js(data["squads"])
    matches_js = matches_to_js(data["matches"])
    users_js   = users_to_js(data.get("users", []))

    html_changed = inject_into_html(squads_js, matches_js, users_js)
    xlsx_changed = generate_xlsx(data.get("users", []))

    if data_changed:
        save_data(data)

    if not any([data_changed, html_changed, xlsx_changed]):
        print("Nothing to do.")
        return 0

    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
